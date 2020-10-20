'''This is a script to automate the monthly MCI Report'''
import pandas as pd, datetime as dt, numpy as np, os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def c_rank(case):
    '''Return the rank of the case type'''
    #Add a column to rank the Case Types 1-5
    caserank = {'Family Investigation':5, 'Treatment':4, 'Permanency':3, 'Guardianship':2, 'Adoption':1}
    return caserank[case]

def s_rank(status):
    '''Return the rank of the case status'''
    statusrank = {'Open':1, 'Closed':2, 'Abridged':3}
    return statusrank[status]

def enterhightlight(x):
    '''Check to see if the Custody Start Date was two months ago'''
    #get last month
    today = dt.datetime.now()
    past = dt.timedelta(days=20)
    lastmonth = today-past
    month = lastmonth.month
    color = '#b3e5fc'
    check = x['Person: Custody Start Date'] < dt.datetime(2020,month,1)
    df1 = pd.DataFrame('', index=x.index, columns=x.columns)
    df1['Person: Custody Start Date'] = np.where(check, 'background-color:{}'.format(color),df1['Person: Custody Start Date'])
    return df1

def exithightlight(x):
    '''Check to see if the Custody Exit Date was two months ago'''
    #get last month
    today = dt.datetime.now()
    past = dt.timedelta(days=20)
    lastmonth = today-past
    month = lastmonth.month
    color = '#b3e5fc'
    check = x['Person: Custody End Date'] < dt.datetime(2020,month,1)
    df1 = pd.DataFrame('', index=x.index, columns=x.columns)
    df1['Person: Custody End Date'] = np.where(check, 'background-color:{}'.format(color),df1['Person: Custody End Date'])
    return df1


def column_size(sheet):
    '''Dynamically adjust the column sizes in excel sheet'''
    column_widths = []
    for row in sheet:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell.value)) > column_widths[i]:
                    column_widths[i] = len(str(cell.value))+5
            else:
                column_widths += [len(str(cell.value))+5]
    for i, column_width in enumerate(column_widths):
        sheet.column_dimensions[get_column_letter(i+1)].width = column_width		

#Ask for file path for both Enter and Exit files and check to see if it ends with xls because that need to be read as an html.    
enterpath = input('Enter the full file path for the MCI Enter download with the extension at the end: \n')
if enterpath.endswith('.xls'):
    enterdf = pd.read_html(enterpath)
    enterdf = enterdf[0]
else:
    enterdf = pd.read_excel(enterpath)

exitpath = input('Enter the full file path for the MCI Exit download with the extension at the end: \n')
if exitpath.endswith('.xls'):
    exitdf = pd.read_html(exitpath)
    exitdf = exitdf[0]
else:
    exitdf = pd.read_excel(exitpath)

kidspath = input('Enter the full file path for the All Highmark kids download with the extension at the end: \n')
if kidspath.endswith('.xls'):
    kidsdf = pd.read_html(kidspath)
    kidsdf = kidsdf[0]
else:
    kidsdf = pd.read_excel(kidspath)

#Ask for date range
drange = input('What is the date range for this report? (ex: 8.1.20 - 9.30.20)\n')

#Get previous months excel "Dups Removed" sheet
prevpath = input('Enter the full file path for the previous MCI Enter report including extension at the end: \n')
preventerdf = pd.read_excel(prevpath, sheet_name="Dups Removed")
prevexitpath = input('Enter the full file path for the previous MCI Exit report including extension at the end: \n')
prevexitdf = pd.read_excel(prevexitpath, sheet_name="Dups Removed")

#Save raw data file as an excel for enter, exit, and kids
mainpath = input('Where do you want to save the files? (ex: C:\\Users\\Name\\Documents\\git\\Work\\MCI_report)\n')
mcienter = r'\MCI Enter Raw Data {}.xlsx'.format(drange)
mciexit = r'\MCI Exit Raw Data {}.xlsx'.format(drange)
enddate = drange.split()[-1]                            #Get end of qtr date
mcikids = r'\MCI Highmark All Kids in Custody {}.xlsx'.format(enddate)

enterwriter = pd.ExcelWriter(mainpath+mcienter, engine='openpyxl', datetime_format = 'MM/DD/YYYY')
exitwriter = pd.ExcelWriter(mainpath+mciexit, engine='openpyxl', datetime_format = 'MM/DD/YYYY')
kidswriter = pd.ExcelWriter(mainpath+mcikids, engine='openpyxl', datetime_format = 'MM/DD/YYYY')

enterdf.to_excel(enterwriter, sheet_name = 'Raw Data', index=None)
exitdf.to_excel(exitwriter, sheet_name = 'Raw Data', index=None)
kidsdf.to_excel(kidswriter, sheet_name = 'Raw Data', index=None)

kidswriter.save()

#Find the max rows of the data and remove the last six rows
alldfs = [enterdf, exitdf, kidsdf]
for df in alldfs:
    if not enterpath.endswith('.xls'):
        maxrow, maxcol = df.shape
        begrow = maxrow - 6
        indx = list(range(begrow,maxrow))
        df.drop(index=indx, inplace=True)

#Change data types for Person PID
alldfs = [enterdf, exitdf, kidsdf]
for df in alldfs:
    df.loc[:,'Person PID'] = pd.to_numeric(df['Person PID'])

#Keep rows that have a case type of Family Investigation, Treatment, Permanency, Guardianship, and Adoption.
casetypes = ['Family Investigation', 'Treatment', 'Permanency', 'Guardianship', 'Adoption']
enterdf1 = enterdf[enterdf['Case Type'].isin(casetypes)].copy()
exitdf1 = exitdf[exitdf['Case Type'].isin(casetypes)].copy()
kidsdf1 = kidsdf[kidsdf['Case Type'].isin(casetypes)].copy()

#Add a new column after Case Type for rank value
enterdf1.insert(10,"Case Rank",enterdf1.loc[:,'Case Type'].apply(c_rank) ,True)
exitdf1.insert(9,"Case Rank",exitdf1.loc[:,'Case Type'].apply(c_rank) ,True)
kidsdf1.insert(10,"Case Rank",kidsdf.loc[:,'Case Type'].apply(c_rank) ,True)

#For exitdf add new column after Case Status for rank value
exitdf1.insert(8,"Status Rank",exitdf1.loc[:,'Case Status'].apply(s_rank) ,True)

#Change Case Open Date and Person: Custody Start and End Date to datetime format
enterdf1.loc[:,['Case Open Date', 'Person: Custody Start Date']] = enterdf1[['Case Open Date', 'Person: Custody Start Date']].apply(pd.to_datetime)
kidsdf1.loc[:,'Case Open Date'] = pd.to_datetime(kidsdf1['Case Open Date'])
exitdf1.loc[:,['Case Open Date', 'Person: Custody End Date']] = exitdf1[['Case Open Date', 'Person: Custody End Date']].apply(pd.to_datetime)

#Sort enter data by Person PID, then Status Rank for Exit file,then Case Rank, then Case Open Date. Then remove duplicates keeping the first.
enterdf1 = enterdf1.sort_values(by=['Person PID', 'Case Rank', 'Case Open Date'], ascending=[True, True, False]).copy()
enterdf1 = enterdf1.drop_duplicates(subset='Person PID', keep='first').copy()

kidsdf1 = kidsdf1.sort_values(by=['Person PID', 'Case Rank', 'Case Open Date'], ascending=[True, True, False]).copy()
kidsdf1 = kidsdf1.drop_duplicates(subset='Person PID', keep='first').copy()
                       
exitdf1 = exitdf1.sort_values(by=['Person PID','Status Rank', 'Case Rank', 'Case Open Date'], ascending=[True, True,True ,False]).copy()
exitdf1 = exitdf1.drop_duplicates(subset='Person PID', keep='first').copy()

#Delete the Case Rank Column and Status Rank Column from exit
dfs = [enterdf1, exitdf1, kidsdf1] 
for df in dfs:
    del df['Case Rank']

del exitdf1['Status Rank']

#Keep rows that were not in previous report. 
preventerpid = preventerdf['Person PID'].copy()
enterdf1 = enterdf1[~enterdf1['Person PID'].isin(preventerpid)].copy()

prevexitpid = prevexitdf['Person PID'].copy()
exitdf1 = exitdf1[~exitdf1['Person PID'].isin(prevexitpid)].copy()

#Sort by Custody Start Date for enterdf and End Date for exitdf
enterdf1 = enterdf1.sort_values(by=['Person: Custody Start Date']).copy()
exitdf1 = exitdf1.sort_values(by=['Person: Custody End Date']).copy()

#Highlight the cell where the the Custody Start Date and End Date for exitdf is two months ago
fin_enterdf = enterdf1.style.apply(enterhightlight, axis=None)
fin_exitdf = exitdf1.style.apply(exithightlight, axis=None)

#Save Dups Removed Sheet into existing excel workbook

fin_enterdf.to_excel(enterwriter, sheet_name = 'Dups Removed', index=None)
enterwriter.save()


fin_exitdf.to_excel(exitwriter, sheet_name = 'Dups Removed', index=None)
exitwriter.save()

#Get the full directory name from file
#dirname = os.path.dirname(mainpath)

#Seperate Insurance Companies to seperate workbooks
highenterdf = enterdf1[enterdf1['Person: Client Person Type: Medicaid MCO']=='Highmark BCBSD Health Options Inc.'].copy()
amerenterdf = enterdf1[enterdf1['Person: Client Person Type: Medicaid MCO']=='Amerihealth Caritas Delaware, Inc'].copy()
fin_highenterdf = highenterdf.style.apply(enterhightlight, axis=None)
fin_amerenterdf = amerenterdf.style.apply(enterhightlight, axis=None)

highexitdf = exitdf1[exitdf1['Person: Client Person Type: Medicaid MCO']=='Highmark BCBSD Health Options Inc.'].copy()
amerexitdf = exitdf1[exitdf1['Person: Client Person Type: Medicaid MCO']=='Amerihealth Caritas Delaware, Inc'].copy()
fin_highexitdf = highexitdf.style.apply(exithightlight, axis=None)
fin_amerexitdf = amerexitdf.style.apply(exithightlight, axis=None)

highpath = r'{}\MCI Highmark {}.xlsx'.format(mainpath,drange)
amerpath = r'{}\MCI Amerihealth {}.xlsx'.format(mainpath,drange)
highwriter = pd.ExcelWriter(highpath, engine='openpyxl', datetime_format = 'MM/DD/YYYY')
amerwriter = pd.ExcelWriter(amerpath, engine='openpyxl', datetime_format = 'MM/DD/YYYY')

fin_highenterdf.to_excel(highwriter, sheet_name = 'Highmark Enter', index=None)
fin_amerenterdf.to_excel(amerwriter, sheet_name = 'Amerhealth Enter', index=None)
fin_highexitdf.to_excel(highwriter, sheet_name = 'Highmark Exit', index=None)
fin_amerexitdf.to_excel(amerwriter, sheet_name = 'Amerhealth Exit', index=None)

kidsdf1.to_excel(highwriter, sheet_name = 'All Highmark Kids', index=None)

highwriter.save()
amerwriter.save()


#Adjust columns sizes
wbs = [mainpath+mcienter,mainpath+mciexit,highpath,amerpath]
for wb in wbs:
    book = load_workbook(wb)
    for sheet in book.sheetnames:
        worksheet = book[sheet]
        column_size(worksheet)
    book.save(wb)               



